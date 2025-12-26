import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Determine base directory based on script location
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, 'individuen.xlsx')

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Individuen"
    
    # Define headers
    headers = [
        "Sarkophag_Inventarnummer",
        "Individuum_ID",
        "Bezeichnung", 
        "Geschlecht",
        "Sterbealter",
        "Kategorie",
        "Erhaltung",
        "Anmerkungen"
    ]
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="364E6F", end_color="364E6F", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        
    # Set column widths
    widths = [25, 15, 25, 15, 20, 25, 20, 50]
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    # Add example data (Herennia Rufina)
    example_data = [
        "CAR-S-1845",
        "A",
        "Individuum A",
        "Weiblich",
        "20 Jahre",
        "Körperbestattung",
        "Gut erhalten",
        "Gemahlin Herennia Rufina"
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border

    wb.save(OUTPUT_FILE)
    print(f"Template erstellt: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_template()
