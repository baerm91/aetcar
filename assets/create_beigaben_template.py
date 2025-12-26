"""
Erstellt eine Excel-Vorlage für Beigaben zu Sarkophagen.
Führe dieses Skript einmal aus, um die Vorlage zu erstellen.
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("openpyxl nicht installiert. Installiere mit: pip install openpyxl")
    sys.exit(1)

# Determine base directory based on script location
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, '..', 'data.json')
OUTPUT_FILE = os.path.join(BASE_DIR, 'beigaben.xlsx')

# Lade Sarkophag-Inventarnummern aus data.json
def get_sarkophag_inventarnummern():
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return sorted([obj.get('Inventarnummer', '') for obj in data if obj.get('Inventarnummer')])
    except Exception as e:
        print(f"Warnung: Konnte data.json nicht laden ({DATA_FILE}): {e}")
        return []

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Beigaben"
    
    # Spaltenüberschriften
    headers = [
        ("Sarkophag_Inventarnummer", "Inventarnummer des Sarkophags (z.B. CAR-S-1845)"),
        ("Beigabe_ID", "Eindeutige ID der Beigabe (z.B. CAR-S-1845-B001)"),
        ("Titel", "Bezeichnung der Beigabe"),
        ("Kategorie", "Münze, Keramik, Schmuck, Glas, Metall, Knochen, Sonstiges"),
        ("Beschreibung", "Kurze Beschreibung der Beigabe"),
        ("Bild_URL", "Link zum Objektfoto"),
        ("Bild_URL_Rueckseite", "Zweiter Link (z.B. für Münz-Rückseite)"),
        ("Emuseum_URL", "Link zur eMuseum-Detailseite"),
        ("Datierung", "Zeitliche Einordnung"),
        ("Material", "Material der Beigabe"),
        ("Masse", "Maße/Gewicht"),
        ("Fundlage", "Position im Sarkophag"),
        ("Bemerkungen", "Zusätzliche Anmerkungen")
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header schreiben
    for col, (header, tooltip) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # Kommentar als Hilfe
        cell.comment = None  # openpyxl Comment könnte hier hinzugefügt werden
    
    # Spaltenbreiten anpassen
    column_widths = [25, 20, 30, 15, 40, 50, 50, 50, 15, 20, 15, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Dropdown für Kategorie (Spalte D)
    kategorie_validation = DataValidation(
        type="list",
        formula1='"Münze,Keramik,Schmuck,Glas,Metall,Knochen,Textil,Sonstiges"',
        allow_blank=True
    )
    kategorie_validation.error = "Bitte eine gültige Kategorie wählen"
    kategorie_validation.errorTitle = "Ungültige Kategorie"
    ws.add_data_validation(kategorie_validation)
    kategorie_validation.add('D2:D1000')
    
    # Sarkophag-Inventarnummern für Dropdown laden
    inventarnummern = get_sarkophag_inventarnummern()
    
    # Zweites Tabellenblatt für Sarkophag-Liste (als Referenz)
    ws_ref = wb.create_sheet("Sarkophage_Referenz")
    ws_ref.cell(row=1, column=1, value="Inventarnummer").font = Font(bold=True)
    for i, inv in enumerate(inventarnummern, 2):
        ws_ref.cell(row=i, column=1, value=inv)
    ws_ref.column_dimensions['A'].width = 25
    
    # Benannten Bereich für Inventarnummern erstellen
    if inventarnummern:
        # Bereich: Sarkophage_Referenz!$A$2:$A$<letzte Zeile>
        last_row = len(inventarnummern) + 1
        ref_range = f"Sarkophage_Referenz!$A$2:$A${last_row}"
        defined_name = DefinedName("Sarkophag_Liste", attr_text=ref_range)
        wb.defined_names.add(defined_name)
        
        # Dropdown für Sarkophag-Inventarnummer mit Referenz auf benannten Bereich
        sarkophag_validation = DataValidation(
            type="list",
            formula1="=Sarkophag_Liste",
            allow_blank=False
        )
        sarkophag_validation.error = "Bitte eine gueltige Inventarnummer waehlen"
        sarkophag_validation.errorTitle = "Ungueltige Inventarnummer"
        ws.add_data_validation(sarkophag_validation)
        sarkophag_validation.add('A2:A1000')
    
    # Beispielzeile
    example_data = [
        "CAR-S-1845",
        "CAR-S-1845-B001",
        "Bronzemünze des Antoninus Pius",
        "Münze",
        "Sesterz mit Porträt des Kaisers",
        "https://example.com/muenze_vorne.jpg",
        "https://example.com/muenze_hinten.jpg",
        "https://emuseum.example.com/object/12345",
        "138-161 n. Chr.",
        "Bronze",
        "Ø 32mm, 25g",
        "Kopfbereich",
        "Gut erhalten"
    ]
    
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = Font(italic=True, color="808080")
        cell.border = thin_border
    
    # Speichern
    wb.save(OUTPUT_FILE)
    print(f"[OK] Excel-Vorlage erstellt: {OUTPUT_FILE}")
    print(f"  - {len(inventarnummern)} Sarkophag-Inventarnummern geladen")
    print(f"  - Beispielzeile eingefügt (bitte überschreiben)")
    print(f"\nNächster Schritt: Beigaben in Excel eintragen, dann convert_beigaben.bat ausführen")

if __name__ == "__main__":
    create_template()
